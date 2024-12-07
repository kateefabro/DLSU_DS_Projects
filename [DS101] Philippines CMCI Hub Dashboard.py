import dash
import plotly.express as px
import dash_bootstrap_components as dbc
from dash import dcc, html, Input, Output
from openpyxl import load_workbook
from pathlib import Path
from dash import Dash, html, dcc, callback, Output, Input
import pandas as pd
import geopandas as gpd
import numpy as npth
import plotly.graph_objects as go
import numpy as np
import copy




# Load datasets
dataset_folder = Path('Datasets/')
workbook_LGU = load_workbook(dataset_folder / 'LGU_Data/LGUs.xlsx')




pillar_data_LGU = {}
pillar_data_PROV = {}




for sheet in workbook_LGU:
    LGUs = []
    scores = []
    distances_km = []
    distances_mi = []
    categories = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if len(row) >= 13:  
            LGUs.append(row[0])
            scores.append(row[1:11])
            distances_km.append(row[11])
            distances_mi.append(row[12])
            categories.append(row[13])  




    pillar_name = sheet.title
    pillar_data_LGU[pillar_name] = {
        'LGUs': LGUs,
        'scores': scores,
        'distances_km': distances_km,
        'distances_mi': distances_mi,
        'categories': categories
    }




workbook_PROV = load_workbook(dataset_folder / 'Province_Data/Prov Dataset.xlsx')




for sheet in workbook_PROV:
    provinces = []
    scores = []
    distances_km = []
    distances_mi = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if len(row) >= 12:  
            provinces.append(row[0])
            scores.append(row[3:13])
            distances_km.append(row[1])
            distances_mi.append(row[2])




    pillar_name = sheet.title
    pillar_data_PROV[pillar_name] = {
        'provinces': provinces,
        'scores': scores,
        'distances_km': distances_km,
        'distances_mi': distances_mi
    }




pillar_names = list(pillar_data_LGU.keys())
pillar_names = ['Overall Score','Economic Dynamism',
                'Government Efficiency',
                'Infrastructure',
                'Innovation',
                'Resiliency']




all_years = list(range(2014, 2024))




LGUs = list(set(LGU for pillar in pillar_names for LGU in pillar_data_LGU[pillar]['LGUs']))




pillar_descriptions = {
    'Overall Score': {
        'Description': 'The sum of scores on five main pillars which pool data from several sub-indicators. The five main pillars are government efficiency, infrastructure, resiliency, economic dynamism, and innovation (innovation was only added starting 2022). Scores are determined by the values of the actual data, as well as the completeness of the submitted data.  The higher the score of a city or municipality, the more competitive it is.'
    },
    'Economic Dynamism': {
        'Description': 'Refers to stable expansion of businesses and industries and higher employment. Matches output and productivity of the local economy with the local resources. Localities are centers of economic activities, and due to this, business expansion and job creation are easily observable in local settings.'
    },
    'Government Efficiency': {
        'Description': 'Refers to the quality and reliability of government services and government support for effective and sustainable productive expansion. This factor looks at government as an institution that is generally not corrupt; able to protect and enforce contracts; apply moderate and reasonable taxation and is able to regulate proactively.'
    },
    'Infrastructure': {
        'Description': 'Refers to the physical assets that connect, expand, and sustain a locality and its surroundings to enable provision of goods and services. It involves basic inputs of production such as energy, water; transportation, roads and communications; sustenance of production such as waste, disaster preparedness, environmental sustainability; and human capital formation infrastructure.'
    },
    'Resiliency': {
        'Description': 'Applies to the capacity of a locality to build systems that can absorb change and disturbance and being able to adapt to such changes. It spans frameworks that bind LGUs and their constituents to prepare for possible shocks and stresses; budgeting for disaster risk reduction; hazard/risk identification mechanisms; resilience-related infrastructure; and resilience-related mechanisms.'
    },
    'Innovation': {
        'Description': 'Refers to the ability of a locality to harness its creative potential to improve or sustain current levels of productivity. It hinges mainly on the development of creative capital which are human resources, research capabilities, and networking capacities. Innovation was only added starting 2022.'
    }
}




pillar_indicators = {
    'Economic Dynamism': [
        '1. Local Economy Size (as measured through business registrations, capital, revenue, and permits)',
        '2. Local Economy Growth (as measured through business registrations, capital, revenue, and permits)',
        '3. Active Establishments in the Locality',
        '4. Safety Compliant Business',
        '5. Employment Generation',
        '6. Cost of Living',
        '7. Cost of Doing Business',
        '8. Financial Deepening',
        '9. Productivity',
        '10. Presence of Business and Professional Organizations'
    ],
    'Government Efficiency': [
        '1. Compliance to National Directives',
        '2. Presence of Investment Promotions Unit',
        '3. Compliance to ARTA Citizens Charter',
        '4. Capacity to Generate Local Resource',
        '5. Capacity of Health Services',
        '6. Capacity of School Services',
        '7. Recognition of Performance',
        '8. Getting Business Permits',
        '9. Peace and Order',
        '10. Social Protection'
    ],
    'Infrastructure': [
        '1. Road Network',
        '2. Distance to Ports',
        '3. Availability of Basic Utilities',
        '4. Transportation Vehicles',
        '5. Education',
        '6. Health',
        '7. LGU Investment',
        '8. Accommodation Capacity',
        '9. Information Technology Capacity',
        '10. Financial Technology Capacity'
    ],
    'Resiliency': [
        '1. Land Use Plan',
        '2. Disaster Risk Reduction Plan',
        '3. Annual Disaster Drill',
        '4. Early Warning System',
        '5. Budget for DRRMP',
        '6. Local Risk Assessments',
        '7. Emergency Infrastructure',
        '8. Utilities',
        '9. Employed Population',
        '10. Sanitary System'
    ],
    'Innovation': [
        '1. Start Up and Innovation Facilities',
        '2. Innovation Financing: R&D Expenditures Allotment',
        '3. Number of STEM graduates',
        '4.Intellectual Property Registration',
        '5. ICT Use: E-BPLS Software',
        '6. Internet Capability',
        '7. Availability of Basic Internet Service',
        '8. Online Payment Facilities',
        '9. ICT Plan',
        '10. New Technology'
    ],
    'Overall Score': [
        'The CMCI has a total index value of 100, representing a fully competitive local unit. The index is composed of five core components with each component representing about 20% of index value.',
        '1. Economic Dynamism - 20%',
        '2. Governance Efficiency - 20%',
        '3. Infrastructure - 20%',
        '4. Resiliency - 20%',
        '5. Innovation - 20%'
    ]
}




pillar_images = {
    'Overall Score': 'https://i.ibb.co/smhd96w/overall-score-3.png',
    'Economic Dynamism': 'https://i.ibb.co/hWH7fsX/economic-dynamism.png',
    'Government Efficiency': 'https://i.ibb.co/q5bc8RC/government-efficiency.png',
    'Infrastructure': 'https://i.ibb.co/6bFxX3H/infrastructure.png',
    'Resiliency': 'https://i.ibb.co/M5GWyrt/resiliency.png',
    'Innovation': 'https://i.ibb.co/BGcySdJ/innovation.png'
}




external_stylesheets = ['https://codepen.io/chriddyp/pen/bWLwgP.css']




workbook = load_workbook(dataset_folder / 'InteractiveMap_Data/InteractiveMap_Profile.xlsx')




# Map
file_paths = [
    dataset_folder / "InteractiveMap_Data/Province JSON/provinces-region-bicolregionregionv.json",
    dataset_folder / "InteractiveMap_Data/Province JSON/provinces-region-autonomousregionofmuslimmindanaoarmm.json",
    dataset_folder / "InteractiveMap_Data/Province JSON/provinces-region-cagayanvalleyregionii.json",
    dataset_folder / "InteractiveMap_Data/Province JSON/provinces-region-calabarzonregioniva.json",
    dataset_folder / "InteractiveMap_Data/Province JSON/provinces-region-caragaregionxiii.json",
    dataset_folder / "InteractiveMap_Data/Province JSON/provinces-region-centralluzonregioniii.json",
    dataset_folder / "InteractiveMap_Data/Province JSON/provinces-region-centralvisayasregionvii.json",
    dataset_folder / "InteractiveMap_Data/Province JSON/provinces-region-cordilleraadministrativeregioncar.json",
    dataset_folder / "InteractiveMap_Data/Province JSON/provinces-region-davaoregionregionxi.json",
    dataset_folder / "InteractiveMap_Data/Province JSON/provinces-region-easternvisayasregionviii.json",
    dataset_folder / "InteractiveMap_Data/Province JSON/provinces-region-ilocosregionregioni.json",
    dataset_folder / "InteractiveMap_Data/Province JSON/provinces-region-metropolitanmanila.json",
    dataset_folder / "InteractiveMap_Data/Province JSON/provinces-region-mimaroparegionivb.json",
    dataset_folder / "InteractiveMap_Data/Province JSON/provinces-region-northernmindanaoregionx.json",
    dataset_folder / "InteractiveMap_Data/Province JSON/provinces-region-soccsksargenregionxii.json",
    dataset_folder / "InteractiveMap_Data/Province JSON/provinces-region-westernvisayasregionvi.json",
    dataset_folder / "InteractiveMap_Data/Province JSON/provinces-region-zamboangapeninsularegionix.json"
]








# Read each file and append to a list
ph_list = [gpd.read_file(file) for file in file_paths]




# Combine all GeoDataFrames in the list into a single GeoDataFrame
ph = gpd.GeoDataFrame(pd.concat(ph_list, ignore_index=True), crs=ph_list[0].crs)
#ph = ph.to_crs(epsg=32651)
p_score = pd.read_csv(dataset_folder / 'Province_Data/Overall Score.csv', encoding='latin1')
ph.loc[ph['PROVINCE'] == 'Agusan del Norte', 'PROVINCE'] = 'Agusan Del Norte'
ph.loc[ph['PROVINCE'] == 'Agusan del Sur', 'PROVINCE'] = 'Agusan Del Sur'
ph.loc[ph['PROVINCE'] == 'Batangas', 'PROVINCE'] = 'Batangas Province'
ph.loc[ph['PROVINCE'] == 'Biliran', 'PROVINCE'] = 'Biliran Province'
ph.loc[ph['PROVINCE'] == 'Cavite', 'PROVINCE'] = 'Cavite Province'
ph.loc[ph['PROVINCE'] == 'Cebu', 'PROVINCE'] = 'Cebu Province'
ph.loc[ph['PROVINCE'] == 'North Cotabato', 'PROVINCE'] = 'Cotabato (North Cotabato)'
ph.loc[ph['PROVINCE'] == 'Davao de Oro', 'PROVINCE'] = 'Davao De Oro'
ph.loc[ph['PROVINCE'] == 'Davao del Norte', 'PROVINCE'] = 'Davao Del Norte'
ph.loc[ph['PROVINCE'] == 'Davao del Sur', 'PROVINCE'] = 'Davao Del Sur'
ph.loc[ph['PROVINCE'] == 'Iloilo', 'PROVINCE'] = 'Iloilo Province'
ph.loc[ph['PROVINCE'] == 'Lanao del Norte', 'PROVINCE'] = 'Lanao Del Norte'
ph.loc[ph['PROVINCE'] == 'Lanao del Sur', 'PROVINCE'] = 'Lanao Del Sur'
ph.loc[ph['PROVINCE'] == 'Leyte', 'PROVINCE'] = 'Leyte Province'
ph.loc[ph['PROVINCE'] == 'Masbate', 'PROVINCE'] = 'Masbate Province'
ph.loc[ph['PROVINCE'] == 'Romblon', 'PROVINCE'] = 'Romblon Province'
ph.loc[ph['PROVINCE'] == 'Samar', 'PROVINCE'] = 'Samar (Western Samar)'
ph.loc[ph['PROVINCE'] == 'Siquijor', 'PROVINCE'] = 'Siquijor Province'
ph.loc[ph['PROVINCE'] == 'Sorsogon', 'PROVINCE'] = 'Sorsogon Province'
ph.loc[ph['PROVINCE'] == 'Surigao del Norte', 'PROVINCE'] = 'Surigao Del Norte'
ph.loc[ph['PROVINCE'] == 'Surigao del Sur', 'PROVINCE'] = 'Surigao Del Sur'
ph.loc[ph['PROVINCE'] == 'Tarlac', 'PROVINCE'] = 'Tarlac Province'
ph.loc[ph['PROVINCE'] == 'Zamboanga del Norte', 'PROVINCE'] = 'Zamboanga Del Norte'
ph.loc[ph['PROVINCE'] == 'Zamboanga del Sur', 'PROVINCE'] = 'Zamboanga Del Sur'
#ph.loc[ph['PROVINCE'] == 'Maguindanao del Norte', 'PROVINCE'] = 'Maguindanao'
#ph.loc[ph['PROVINCE'] == 'Maguindanao del Sur', 'PROVINCE'] = 'Maguindanao'
ph.loc[ph['PROVINCE'] == 'Metropolitan Manila', 'PROVINCE'] = 'Metro Manila'
p_choro = pd.merge(ph, p_score,  left_on='PROVINCE', right_on='PROVINCE / LGU', how='left', indicator=True)




province_options = [{'label': province, 'value': province} for province in p_choro['PROVINCE'] if province is not None]




# Province profile
province_sheet = workbook['Province']




province = []
region = []
population = []
province_revenue = []
rank = []




for row in province_sheet.iter_rows(min_row=2, values_only=True):
   region.append(row[1])
   population.append(row[2])
   province_revenue.append(row[3])
   rank.append(row[4])




province_data = [row[0].value for row in province_sheet.iter_rows(min_row=2)]
region_data = [row[1].value for row in province_sheet.iter_rows(min_row=2)]
population_data = [row[2].value for row in province_sheet.iter_rows(min_row=2)]
province_revenue_data = [row[3].value for row in province_sheet.iter_rows(min_row=2)]
rank_data = [row[4].value for row in province_sheet.iter_rows(min_row=2)]




def get_province_region(province):
   try:
       index = province_data.index(province)
       return region_data[index]
   except ValueError:
       return 'No data available'
   
def get_province_population(province):
   try:
       index = province_data.index(province)
       return population_data[index]
   except ValueError:
       return 'No data available'








def get_province_revenue(province):
   try:
       index = province_data.index(province)
       return province_revenue[index]
   except ValueError:
       return 'No data available'








def get_province_rank(province):
   try:
       index = province_data.index(province)
       return rank_data[index]
   except ValueError:
       return 'No data available'
   
# LGU Profile
lgu_sheet = workbook['LGU']




lgu = []
category = []
percentage = []
lgu_province = []
revenue = []




for row in lgu_sheet.iter_rows(min_row=2, values_only=True):
   lgu.append(row[0])
   category.append(row[1])
   percentage.append(row[2])
   lgu_province.append(row[3])
   revenue.append(row[4])




lgu_data = [row[0].value for row in lgu_sheet.iter_rows(min_row=2, max_col=1)]
category_data = [row[1].value for row in lgu_sheet.iter_rows(min_row=2, max_col=2)]
lgu_province_data = [row[3].value for row in lgu_sheet.iter_rows(min_row=2, max_col=4)]
revenue_data = [row[4].value for row in lgu_sheet.iter_rows(min_row=2, max_col=5)]
lgu_options = [{'label': lgu_name, 'value': lgu_name} for lgu_name in lgu_data if lgu_name is not None]




def get_lgu_province(selected_lgu):
   try:
       index = lgu_data.index(selected_lgu)
       return lgu_province_data[index]
   except ValueError:
       return 'No data available'




def get_lgu_category(selected_lgu):
   try:
       index = lgu_data.index(selected_lgu)
       return category_data[index]
   except ValueError:
       return 'No data available'




def get_lgu_revenue(selected_lgu):
   try:
       index = lgu_data.index(selected_lgu)
       return revenue_data[index]
   except ValueError:
       return 'No data available'




app = dash.Dash(__name__, external_stylesheets=[dbc.themes.BOOTSTRAP],suppress_callback_exceptions=True)








page1_layout = html.Div(
    style={'background-image': 'url(assets/phmap1.png)', 'height': '100vh','background-repeat': 'no-repeat','background-size': 'cover', 'position': 'relative'},
    children=[
        html.Div(
            style={'display': 'flex', 'flex-direction': 'column', 'justify-content': 'center', 'align-items': 'left', 'height': '100%', 'margin-left':'20px'},
            children=[
                html.Div(
                    style={'text-align': 'left'},
                    children=[
                        html.H1('Explore CMCI Data with Ease',style={'font-size':'70px', 'color': '#000033'}),
                        html.Hr(style={'border-top': '3px solid navy', 'width':'50%'}),
                        html.P(
                            'View DTI’s Rankings of Cities and Municipalities. The overall score is based on the sum of their scores on',
                            style={'text-align':'justify','color': '#000033'},
                        ),
                        html.P(
                            '5 Pillars: Economic Dynamism, Government Efficiency, Infrastructure, Innovation, and Resiliency.',
                            style={'text-align':'left','color': '#000033'},
                        )
                    ]
                ),
            ]
        )
    ]
)
available_years = [2014, 2015, 2016, 2017, 2018, 2019, 2020, 2021, 2022, 2023]
year_options = [{'label': str(year), 'value': year} for year in available_years]








page2_layout = dbc.Container([
    # Header
    html.Div([
        html.H1("Dashboard", style={'text-align': 'left', 'font-size': '50px', 'margin-top':'20px'}),
        html.Div([
            # Level Dropdown
            html.Label('Level',style={'margin-top':'10px'}),
            dcc.Dropdown(
                id='level-dropdown',
                options=[
                    {'label': 'LGU', 'value': 'LGU'},
                    {'label': 'Province', 'value': 'Province'}
                ],
                value='LGU',  
                style={'width': '100px', 'display': 'inline-block'}
            ),
            # Pillar Dropdown
            html.Label('Pillar',style={'margin-top':'10px'}),
            dcc.Dropdown(
                id='pillar-dropdown',
                options=[{'label': pillar, 'value': pillar} for pillar in pillar_names],
                value='Overall Score',
                style={'width': '200px', 'display': 'inline-block'}
            ),
            # Starting Year Dropdown
            html.Label('Starting Year',style={'margin-top':'10px'}),
            dcc.Dropdown(
                id='start-year-dropdown',
                options=[{'label': str(year), 'value': year} for year in all_years],
                value=2014,  
                style={'width': '80px', 'display': 'inline-block'}
            ),
            # End Year Dropdown
            html.Label('End Year',style={'margin-top':'10px'}),
            dcc.Dropdown(
                id='end-year-dropdown',
                options=[{'label': str(year), 'value': year} for year in all_years],
                value=2023,  
                style={'width': '80px', 'display': 'inline-block'}
            ),
        ], style={'display': 'flex', 'flex-direction': 'row', 'gap': '20px', 'padding': '20px', 'width':'100%','margin-right':'20px'}),
    ], id='header', style={'display': 'flex', 'flex-direction': 'row', 'text-align': 'center'}),




    # Row 2 (Displayed conditionally)
    dbc.Row([
        # First column
        dbc.Col([
            dbc.Card([
                dbc.CardBody([
                    html.Div(id='pillar-info-container')
                ])
            ], color='light', style={'margin-bottom': '20px'})
        ], id='row2-col1', width=3),
        # Second column
        dbc.Col([
            dbc.Card([
                dbc.CardBody([
                    html.Div([
                        html.H3('Scores over Time', style={'text-align': 'center'}),
                        dcc.Graph(id='line-chart')
                    ])
                ])
            ], color='light', style={'margin-bottom': '20px'})
        ], id='row2-col2', width=6),
   




        # Third column
        dbc.Col([
                dbc.Card([
                    dbc.CardBody([
                        html.Div([
                            html.H4('Select LGUs', style={'display': 'block', 'text-align': 'center', 'margin-bottom': '5px'}),
                            # Search bar
                            dcc.Input(id='LGU-search', type='text', placeholder='Search LGUs...'),
                            dcc.Checklist(
                                id='LGU-checkboxes',
                                options=[{'label': LGU, 'value': LGU} for LGU in LGUs],
                                value=[],
                                style={'overflowY': 'scroll', 'height': '400px'}
                            ),
                            html.Button('Clear Selection', id='clear-selection-button', n_clicks=0)
                        ], style={'margin-left': '20px'})
                    ])
                ], color='light', style={'margin-bottom': '20px'})
            ], id='row2-col3', width=3)
    ], id='row2', style={'display': 'none'}),




    # Row 3 (Displayed conditionally)
    dbc.Row([
        # First column
        dbc.Col([
            dbc.Card([
            dbc.CardBody([
                    html.H3('Pillar Indicators', style={'text-align': 'center'}),
                    html.Div(id='pillar-indicators-container')
                ])
            ], color='light', style={'margin-bottom': '20px'})
        ], id='row3-col1', width=3),
        # Second column
        dbc.Col([
            dbc.Card([
                dbc.CardBody([
                    html.Div([
                        html.H3('Composition of Overall Score', style={'text-align': 'center'}),
                        html.Div([
                            html.Label('Select Year', style={'display': 'inline-block', 'vertical-align': 'middle', 'margin-right': '10px'}),
                            dcc.Dropdown(
                                id='bar-year-dropdown-lgu',
                                options=[{'label': str(year), 'value': year} for year in all_years],
                                value=2023,
                                style={'width': '80px', 'margin-bottom': '1px', 'display': 'inline-block', 'vertical-align': 'middle'}
                            ),
                        ], style={'text-align': 'center'}),
                        dcc.Graph(id='bar-chart'),
                    ])
                ])
            ], color='light', style={'margin-bottom': '20px'})
        ], id='row3-col2', width=6),
        # Third column
        dbc.Col([
             dbc.Card([
                dbc.CardBody([
                    html.Div([
                        html.H3('LGU Information', style={'text-align': 'center', 'margin-bottom': '10px'}),
                        html.Div(id='table-container', style={'margin-bottom': '20px'})
                    ])
                ])
            ], color='light', style={'margin-bottom': '20px'})
        ], id='row3-col3', width=3)
    ], id='row3', style={'display': 'none'}),




    # Row 4 (Displayed conditionally)
    dbc.Row([        
    # Column 1
    dbc.Col([
            #Row 1: Line chart
            dbc.Row([
                dbc.Card([
                    dbc.CardBody([
                        html.Div([
                            html.H3('CMCI Scores', style={'text-align': 'center'}),
                            dcc.Graph(id='line-chart-prov'),
                        ]),
                    ])
                ],color='light', style={'margin-bottom': '20px','margin-right':'20px','height':'400px'})
            ]),
            #Row 2: Scatter Chart
            dbc.Row([
                dbc.Card([
                    dbc.CardBody([
                        html.Div([
                            html.H3('Overall Score vs Distance', style={'text-align': 'center'}),
                                dcc.Dropdown(
                                    id='scatter-year-dropdown',
                                    options=year_options,
                                    value=2023,
                                    style={'width': '100%'}  
                                ),
                            dcc.Graph(id='scatter-plot-prov')
                        ])
                    ])
                ],color='light', style={'margin-bottom': '20px', 'margin-right':'25px','height':'600px'})
            ])
        ], width=4, style={'padding-left': '20px'}),




    # Map
    dbc.Col([
            dbc.Card([
                        dbc.CardBody([
                            dcc.Dropdown(
                                id='map-year-dropdown-province-1',
                                options=[{'label': str(year), 'value': year} for year in all_years],
                                value=2023,
                                style={'width': '80px', 'display': 'inline-block', 'vertical-align': 'middle','margin-top': '10px'}
                            ),
                            html.Div([
                                html.H3('Overall CMCI Score per Province', style={'text-align': '-webkit-center',  'margin': '0 0 10px 0','display': 'inline-block','margin-top':'0'}),
                                dcc.Loading(
                                    id="loading-choropleth-map",
                                    type="default",
                                    children=dcc.Graph(id='choropleth-map-1', style={'height': '890px','width': '100%'})
                                )
                    ], style={'text-align': 'center', 'display': 'inline-block','margin': '0 0 10px 0' }),
                        ])
            ], color='light', style={'margin-bottom': '20px','height':'1020px'})
    ], width=5),




    # Sidebar
    dbc.Col([
        dbc.Card([
            dbc.CardBody([
                html.Div([
                    html.Label('Select Provinces', style={'display': 'block', 'text-align': 'center', 'margin-bottom': '5px','font-weight':'bold','font-size':'20px'}),
                    dcc.Input(id='province-search', type='text', placeholder='Search Provinces...'),
                    dcc.Checklist(
                        id='province-checkboxes',
                        options=[{'label': province, 'value': province} for province in provinces],
                        value=[],
                        style={'overflowY': 'scroll', 'height': '890px'}
                    ),
                    html.Button('Clear Selection', id='clear-selection-button-prov', n_clicks=0)
                ], style={'margin-left': '20px'}),
            ]),
        ], color='light', style={'margin-bottom': '20px','height':'1020px'})
    ], width=3),
], id='row4', style={'display': 'none'})




], fluid=True)




page3_layout = dbc.Container([
 
   dbc.Row([
       
   # First column
   dbc.Col([
       # Card 1, Entire First Column
       dbc.Card([
           dbc.CardBody([
           dbc.Row([
                dbc.Row([
                    html.Div([
                    html.H3('Choropleth Map', style={'text-align': 'center', 'margin-bottom': '10px'}),
                    dbc.Row([
                        html.Label('Select Year', style={'font-weight': 'bold'}),
                        dcc.Dropdown(
                            id='map-year-dropdown-province',
                            options=[{'label': str(year), 'value': year} for year in all_years],
                            value=2023,
                            style={'width': '80px', 'margin-bottom': '1px', 'display': 'inline-block', 'vertical-align': 'middle'}
                        )])
                       
                   
                   
                ]),
                    dcc.Loading(
                        id="loading-choropleth-map",
                        type="default",
                        children=dcc.Graph(id='choropleth-map')
                                )
                ])
            ]),
            ], style={'height': '953px'})
       ], color='light')
       
       
   ]),
 
   # Second column
   dbc.Col([
       html.Div([
           # Card 2, Top Right
           dbc.Card([dbc.CardBody([
                dbc.Row([
                html.H3('Province Profile', style={'text-align': 'center'}),
                html.Label('Select Province', style={'font-weight': 'bold'}),
                dcc.Dropdown(
                    id='province-dropdown',
                    options=sorted(province_options, key=lambda d: d['label']),
                    value=[]
                ),
                dbc.Row([
                    dbc.Col([
                        html.Div(id='map_prov_table')
                    ])
                   
                ], justify='center',align='center', style={'margin-bottom': '10px', 'margin-top':'10px'})
                ])
           ], style={'height': '220px'})], color='light'),


           # Card 3, Bottom Right
           dbc.Card([dbc.CardBody([
               dbc.Row([
               html.H3('LGU Profile', style={'text-align': 'center'}),
                # LGU Dropdown
                html.Label('Select LGU', style={'font-weight': 'bold'}),
                dcc.Dropdown(
                    id='lgu-dropdown',
                    options=lgu_options,
                    value=[]
                ),
                dbc.Row([
                    dbc.Col([
                        html.Div(id='map_lgu_table')
                    ])
                ], justify='center',align='center', style={'margin-bottom': '10px', 'margin-top':'10px'}),
                dbc.Row([
                    html.Div([
                        dcc.Graph(id='bar-chart-map')
                    ])
                ]),
                dbc.Row([
                    dbc.Col([
                        html.Label('Select Pillar', style={'font-weight': 'bold'}),
                        dcc.Dropdown(
                            id='pillar-dropdown-map',
                            options=[
                                {'label': 'Resiliency', 'value': 'Resiliency'},
                                {'label': 'Government Efficiency', 'value': 'Government Efficiency'},
                                {'label': 'Innovation', 'value': 'Innovation'},
                                {'label': 'Economic Dynamism', 'value': 'Economic Dynamism'},
                                {'label': 'Infrastructure', 'value': 'Infrastructure'},
                            ],
                            value=[]
                        ),
                    ]),
                    dbc.Col([
                        html.Label('Pillar Description', style={'font-weight': 'bold'}),
                        html.Div(id='pillar-description')
                    ])
                ]),
            ])
           ], style={'height': '720px'})], color='light', style={'margin-top': '10px'})
           
   ]),
   ])
   ])
], style={'margin-top': '20px'})




# Define navigation bar
navbar = dbc.NavbarSimple(
    children=[
        html.Div(
        [
            dbc.Button("🌐 CMCI HUB", href="/page-1", color="secondary", style={"margin-right": "20rem"}),
            dbc.Button("📊 VISUALIZATION DASHBOARD", href="/page-2", color="secondary", style={"margin-right": "20rem"}),
            dbc.Button("🗾 INTERACTIVE MAP", href="/page-3", color="secondary")
        ],




            className="d-flex justify-content-center"
        )
    ],
    color="dark",
    dark=True,  
    style={"font-family": "Arial, sans-serif", "font-weight": "bold", "color": "black"}
)




@app.callback(
    [Output('row2', 'style'), Output('row3', 'style'), Output('row4', 'style')],
    [Input('level-dropdown', 'value')]
)
def update_row_visibility(level):
    row2_style = {'display': 'none'} if level != 'LGU' else {}
    row3_style = {'display': 'none'} if level != 'LGU' else {}
    row4_style = {'display': 'none'} if level != 'Province' else {}
    return row2_style, row3_style, row4_style




@app.callback(
    Output('pillar-dropdown', 'value'),
    [Input('level-dropdown', 'value')]
)
def set_pillar_dropdown(level):
    if level == 'Province' or level == 'LGU':
        return 'Overall Score'




@app.callback(
    Output('province-checkboxes', 'options'),
    [Input('province-search', 'value')]
)
def update_province_options_prov(search_value):
    if search_value is None:
        return [{'label': province, 'value': province} for province in provinces]
    else:
        filtered_provinces = [province for province in provinces if search_value.lower() in province.lower()]
        return [{'label': province, 'value': province} for province in filtered_provinces]




# Callback to clear selected provinces
@app.callback(
    Output('province-checkboxes', 'value'),
    [Input('clear-selection-button-prov', 'n_clicks')]
)
def clear_selected_provinces(n_clicks):
    if n_clicks > 0:
        return []
    else:
        raise dash.exceptions.PreventUpdate




@app.callback(
    Output('line-chart-prov', 'figure'),
    [
        Input('pillar-dropdown', 'value'),
        Input('start-year-dropdown', 'value'),
        Input('end-year-dropdown', 'value'),
        Input('province-checkboxes', 'value'),
    ]
)
def update_data_prov(pillar, start_year, end_year, selected_provinces):
    # Filter data based on selected pillar and years
    selected_data = pillar_data_PROV[pillar]
    filtered_provinces = []
    filtered_scores = []
   
    # Colors
    color_palette = px.colors.qualitative.Plotly
   
    # Filter data for selected provinces
    for index, (province, scores) in enumerate(zip(selected_data['provinces'], selected_data['scores'])):
        if province in selected_provinces:
            filtered_provinces.append(province)
            filtered_scores.append(scores)
   
    if not filtered_provinces:
        no_data_layout = {
            'xaxis': {'visible': False},
            'yaxis': {'visible': False},
            'annotations': [{
                'text': 'No matching data found',
                'xref': 'paper',
                'yref': 'paper',
                'showarrow': False,
                'font': {'size': 28}
        }],
        'height': 300  
    }
        return {'layout': no_data_layout}
   
    else:
        line_chart_data = []
        for province, scores, color in zip(filtered_provinces, filtered_scores, color_palette):
            line_chart_data.append({
                'x': list(range(start_year - 1, end_year)),
                'y': scores,
                'mode': 'lines',
                'name': province,
                'line': {'color': color}
            })




        return {'data': line_chart_data, 'layout': {'title': f'{pillar} scores by Province over Time',
                                                    'xaxis': {'title': 'Year'},
                                                    'yaxis': {'title': 'Score'},
                                                    'height': 300}}




def filter_data_by_year_range(data, selected_year):
    filtered_data = {'provinces': [], 'scores': [], 'distances_mi': []}
   
    for province, scores, distance_mi in zip(data['provinces'], data['scores'], data['distances_mi']):
        if any(selected_year <= year <= selected_year for year in range(2014, 2024)):
            # Add province data to filtered data
            filtered_data['provinces'].append(province)
            filtered_data['scores'].append(scores)
            filtered_data['distances_mi'].append(distance_mi)
   
    return filtered_data




@app.callback(
    Output('scatter-plot-prov', 'figure'),
    [Input('scatter-year-dropdown', 'value')],
)


def update_scatter_plot_prov(selected_year):
    year_index = selected_year - 2014
    filtered_data = copy.deepcopy(pillar_data_PROV['Overall Score'])
    filtered_scores = []
    for score_list in filtered_data['scores']:
        filtered_scores.append(score_list[year_index])
    filtered_data['scores'] = filtered_scores


    all_scores = []
    all_distances_mi = []
    all_provinces = []
   
    for scores, distances_mi, provinces in zip(filtered_data['scores'], filtered_data['distances_mi'], filtered_data['provinces']):
            all_scores.append(scores)  
            all_distances_mi.append(distances_mi)  
            all_provinces.append(provinces)


    marker_colors = ['lightgreen' if distance <= 300 else 'darkgreen' for distance in all_distances_mi]


    scatter_plot_data = [{
            'x': all_distances_mi,
            'y': all_scores,
            'mode': 'markers',
            'marker': {
                'size': 10,
                'opacity': 0.6,
                'color': marker_colors,  
            },
            'text': [f"{province}<br>Distance: {distance} mi<br>Overall Score: {score}"
                    for province, distance, score in zip(all_provinces, all_distances_mi, all_scores)],
            'hoverinfo': 'text',
        }]
   
    scatter_plot_layout = {
            'title': 'Distance of Each Province to the Center of Manila',
            'xaxis': {'title': 'Distance (mi)'},
            'yaxis': {'title': 'Overall Score'},
            'hovermode': 'closest',
            'height': 500
        }


    fig = go.Figure(data=scatter_plot_data, layout=scatter_plot_layout)


    return fig




@app.callback(
    Output('LGU-checkboxes', 'options'),
    [Input('LGU-search', 'value')]
)
def update_LGU_options(search_value):
    if search_value is None:
        sorted_LGUs = sorted(LGUs)
        return [{'label': LGU, 'value': LGU} for LGU in sorted_LGUs]
    else:
        filtered_LGUs = [LGU for LGU in LGUs if search_value.lower() in LGU.lower()]
        sorted_filtered_LGUs = sorted(filtered_LGUs)
        return [{'label': LGU, 'value': LGU} for LGU in sorted_filtered_LGUs]








@app.callback(
    Output('LGU-checkboxes', 'value'),
    [Input('clear-selection-button', 'n_clicks')]
)
def clear_selected_LGUs(n_clicks):
    if n_clicks > 0:
        return []
    else:
        raise dash.exceptions.PreventUpdate




@app.callback(
    [
        Output('table-container', 'children'),
        Output('line-chart', 'figure'),
        Output('bar-chart', 'figure')
    ],
    [
        Input('pillar-dropdown', 'value'),
        Input('start-year-dropdown', 'value'),
        Input('end-year-dropdown', 'value'),
        Input('LGU-checkboxes', 'value'),
        Input('bar-year-dropdown-lgu', 'value')
    ]
)
def update_data(pillar, start_year, end_year, selected_LGUs, bar_chart_year):
    # Check if no LGUs are selected
    if not selected_LGUs:
        no_data_layout = {
            'xaxis': {'visible': False},
            'yaxis': {'visible': False},
            'annotations': [{
                'text': 'Please select LGUs',
                'xref': 'paper',
                'yref': 'paper',
                'showarrow': False,
                'font': {'size': 28}
            }],
            'height': 450
        }
        return '', {'layout': no_data_layout}, {'layout': no_data_layout}


    # Filtered LGUs
    filtered_LGUs = selected_LGUs


    # Table
    table_rows = [
        html.Tr([
            html.Th('LGU', style={'border-bottom': '1px solid #ddd', 'font-size': '14px', 'text-align': 'center', 'background-color': '#F1F1F1'}),
            html.Th('Category', style={'border-bottom': '1px solid #ddd', 'font-size': '14px', 'text-align': 'center', 'background-color': '#F1F1F1'}),
            html.Th('Distance from MNL (km)', style={'border-bottom': '1px solid #ddd', 'font-size': '14px', 'text-align': 'center', 'background-color': '#F1F1F1'}),
            html.Th('Distance from MNL (mi)', style={'border-bottom': '1px solid #ddd', 'font-size': '14px', 'text-align': 'center', 'background-color': '#F1F1F1'})
        ])
    ]


    for i, LGU in enumerate(filtered_LGUs):
        LGU_index = pillar_data_LGU[pillar]['LGUs'].index(LGU)
        distances_km = pillar_data_LGU[pillar]['distances_km'][LGU_index]
        distances_mi = pillar_data_LGU[pillar]['distances_mi'][LGU_index]
        category = pillar_data_LGU[pillar]['categories'][LGU_index]


        if i % 2 == 0:
            row_style = {'background-color': '#F9F7F7'}
        else:
            row_style = {'background-color': 'white'}


        table_rows.append(html.Tr([
            html.Td(LGU, style={'border-bottom': '1px solid #ddd', 'font-size': '14px', 'text-align': 'center', **row_style}),
            html.Td(category, style={'border-bottom': '1px solid #ddd', 'font-size': '14px', 'text-align': 'center', **row_style}),
            html.Td(distances_km, style={'border-bottom': '1px solid #ddd', 'font-size': '14px', 'text-align': 'center', **row_style}),
            html.Td(distances_mi, style={'border-bottom': '1px solid #ddd', 'font-size': '14px', 'text-align': 'center', **row_style})
        ]))


    # Line chart
    line_chart_data = []
    for LGU in filtered_LGUs:
        LGU_index = pillar_data_LGU[pillar]['LGUs'].index(LGU)
        scores = pillar_data_LGU[pillar]['scores'][LGU_index]
        line_chart_data.append({
            'x': list(range(start_year, end_year + 1)),
            'y': scores[start_year - 2014:end_year - 2014 + 1],
            'mode': 'lines',
            'name': LGU,
        })


    # Bar chart
    bar_chart_data = []
    selected_year_index = bar_chart_year - 2014
    filtered_pillar_names = [p for p in pillar_names if p != 'overall score']
    for j, pillar_name in enumerate(filtered_pillar_names):
        pillar_scores = []
        for LGU in filtered_LGUs:
            LGU_index = pillar_data_LGU[pillar_name]['LGUs'].index(LGU)
            scores = pillar_data_LGU[pillar_name]['scores'][LGU_index]
            score = scores[selected_year_index] if scores[selected_year_index] != '-' else 0
            pillar_scores.append(score)


        bar_chart_data.append({
            'x': filtered_LGUs,
            'y': pillar_scores,
            'name': pillar_name,
            'type': 'bar',
        })
    bar_chart_layout = {
        'title': f'Composition of Overall Score for {bar_chart_year}',
        'xaxis': {'title': 'LGU'},
        'yaxis': {'title': 'Score'}
    }
    return html.Table(table_rows), {'data': line_chart_data, 'layout': {'title': f'{pillar} scores by LGU over Time',
                                                                         'xaxis': {'title': 'Year'},
                                                                         'yaxis': {'title': 'Score'}}}, {
               'data': bar_chart_data, 'layout': bar_chart_layout}


@app.callback(
    Output('pillar-info-container', 'children'),
    [Input('pillar-dropdown', 'value')]
)
def update_pillar_info(pillar):
    if pillar in pillar_descriptions:
        description = pillar_descriptions[pillar]['Description']
        image_url = pillar_images.get(pillar, '')
        return html.Div([
            html.H3('Pillar Description', style={'text-align': 'center'}),
            html.H5(f'{pillar.upper()}', style={'text-align': 'center'}),
            html.Img(src=image_url, style={'display': 'block', 'margin': 'auto','width': '50%'}),
            html.P(description, style={'text-align': 'justify'})
        ], style={'margin': 'auto', 'width': '80%', 'height':'100%'})
    else:
        return 'No information available for selected pillar'




@app.callback(
    Output('pillar-indicators-container', 'children'),
    [Input('pillar-dropdown', 'value')]
)
def update_pillar_indicators_table(pillar):
    indicators = pillar_indicators.get(pillar, [])
    table_rows = [html.Tr([html.Td(indicator, style={'text-align': 'justify'})]) for indicator in indicators]
    return html.Table([
        html.Thead(html.Tr([html.H5(pillar.upper())], style={'background-color': '#f9f9f9', 'text-align': 'center'})),
        html.Tbody(table_rows)
    ], style={'width': '100%'})




# Descriptions
# Map Page, Table Province
@app.callback(
    [
        Output('map_prov_table','children')
    ],
    [
        Input('province-dropdown','value')
    ]
)
def update_labels(province):
    if province:
       province_region = get_province_region(province)
       province_population = get_province_population(province)
       province_revenue = get_province_revenue(province)
       province_rank = get_province_rank(province)


       table_rows = [
                    html.Tr([
                            html.Th('Region', style={'border-bottom': '1px solid #ddd', 'font-size': '14px', 'text-align': 'center', 'background-color': '#F1F1F1', 'width': '25%'}),  
                            html.Th('Population', style={'border-bottom': '1px solid #ddd', 'font-size': '14px', 'text-align': 'center', 'background-color': '#F1F1F1', 'width': '25%'}),
                            html.Th('Revenue', style={'border-bottom': '1px solid #ddd', 'font-size': '14px', 'text-align': 'center', 'background-color': '#F1F1F1', 'width': '25%'}),
                            html.Th('Ranking', style={'border-bottom': '1px solid #ddd', 'font-size': '14px', 'text-align': 'center', 'background-color': '#F1F1F1', 'width': '25%'})
                            ], style={'width': '100%'})
                    ]
       table_rows.append(html.Tr([
            html.Td(province_region, style={'border-bottom': '1px solid #ddd', 'font-size': '14px', 'text-align': 'center', 'width': '25%'}),  
            html.Td(province_population, style={'border-bottom': '1px solid #ddd', 'font-size': '14px', 'text-align': 'center', 'width': '25%'}),
            html.Td(province_revenue, style={'border-bottom': '1px solid #ddd', 'font-size': '14px', 'text-align': 'center', 'width': '25%'}),
            html.Td(province_rank, style={'border-bottom': '1px solid #ddd', 'font-size': '14px', 'text-align': 'center', 'width': '25%'})
        ], style={'width': '100%'}))
       return [html.Table(table_rows, style={'width': '100%', 'margin': 'auto', 'textAlign': 'center'})]
    else:
        return [[]]  # Return an empty list to match the expected output type


# Map LGU Table
@app.callback(
    [
        Output('map_lgu_table','children')
    ],
    [
        Input('lgu-dropdown','value')
    ]
)
def update_labels(selected_lgu):
    if selected_lgu:
       lgu_province = get_lgu_province(selected_lgu)
       lgu_category = get_lgu_category(selected_lgu)
       lgu_revenue = get_lgu_revenue(selected_lgu)


       table_rows = [
                    html.Tr([
                            html.Th('Province', style={'border-bottom': '1px solid #ddd', 'font-size': '14px', 'text-align': 'center', 'background-color': '#F1F1F1', 'width': '34%'}),  
                            html.Th('Category', style={'border-bottom': '1px solid #ddd', 'font-size': '14px', 'text-align': 'center', 'background-color': '#F1F1F1', 'width': '33%'}),
                            html.Th('Revenue', style={'border-bottom': '1px solid #ddd', 'font-size': '14px', 'text-align': 'center', 'background-color': '#F1F1F1', 'width': '33%'})
                            ], style={'width': '100%'})
                    ]
       table_rows.append(html.Tr([
            html.Td(lgu_province, style={'border-bottom': '1px solid #ddd', 'font-size': '14px', 'text-align': 'center', 'width': '34%'}),  
            html.Td(lgu_category, style={'border-bottom': '1px solid #ddd', 'font-size': '14px', 'text-align': 'center', 'width': '33%'}),
            html.Td(lgu_revenue, style={'border-bottom': '1px solid #ddd', 'font-size': '14px', 'text-align': 'center', 'width': '33%'})
        ], style={'width': '100%'}))
       return [html.Table(table_rows, style={'width': '100%', 'margin': 'auto', 'textAlign': 'center'})]
    else:
        return [[]]  


def get_pillar_description(selected_pillar):
   pillar_descriptions = {
       'Resiliency': 'Applies to the capacity of a locality to build systems that can absorb change and disturbance and being able to adapt to such changes',
       'Government Efficiency': 'Refers to the quality and reliability of government services and government support for effective and sustainable productive expansion',
       'Innovation': 'Refers to the ability of a locality to harness its creative potential to improve or sustain current levels of productivity',
       'Economic Dynamism': 'Refers to stable expansion of businesses and industries and higher employment',
       'Infrastructure': 'Pertains to the physical assets that connect, expand, and sustain a locality and its surroundings to enable provision of goods and services'
   }
   return pillar_descriptions.get(selected_pillar, 'No description available')


@app.callback(
    Output('pillar-description', 'children'),
    [Input('pillar-dropdown-map', 'value')]
)
def update_pillar_description(selected_pillar):
    if not selected_pillar:
        return "-"
    else:
        return get_pillar_description(selected_pillar)


# Bar Chart
@app.callback(
   Output('bar-chart-map', 'figure'),
   Input('lgu-dropdown', 'value')
)
def update_bar_chart(selected_lgu):
    if not selected_lgu:
        no_data_layout = {
            'xaxis': {'visible': False},
            'yaxis': {'visible': False},
            'annotations': [{
                'text': 'Please select an LGU above.',
                'xref': 'paper',
                'yref': 'paper',
                'showarrow': False,
                'font': {'size': 28}
        }],
        'height': 300  
    }
        return {'layout': no_data_layout}
   
    else:


        lgu_index = lgu_data.index(selected_lgu) + 2
        lgu_data_row = list(lgu_sheet.iter_rows(min_row=lgu_index, max_row=lgu_index, min_col=6, max_col=10, values_only=True))[0]
        pillars = ['Resiliency', 'Government Efficiency', 'Innovation', 'Economic Dynamism', 'Infrastructure']


        fig = px.bar(
            x=pillars,
            y=lgu_data_row,
            labels={'x': 'Pillar', 'y': 'Score'},
            color=pillars,
            height=400,
            title=f'Scores per Pillar for {selected_lgu}'
        )
        fig.update_layout(
                showlegend=False,
                title_font=dict(size=20, family='Arial Black'),
            )
       
        fig.update_traces(hovertemplate='Pillar: %{x} <br>Score: %{y}<extra></extra>')
    return fig


# Map 1
@app.callback(
    Output('choropleth-map-1', 'figure'),
    Input('map-year-dropdown-province-1', 'value')
)
def update_choropleth(map_year):
    initial_column_values = p_choro.set_index('PROVINCE')[str(map_year)].replace('-', np.nan).astype(float).fillna(0)


    initial_fig = px.choropleth_mapbox(
        p_choro,
        geojson=p_choro,
        locations=p_choro.PROVINCE,
        featureidkey="properties.PROVINCE",
        color=initial_column_values,
        color_continuous_scale='Viridis',
        hover_name='PROVINCE',  
        hover_data={str(map_year): True},
        labels={map_year: 'Overall CMCI Score'},
        center={'lat': 12.8797, 'lon': 121.7740},
        mapbox_style="carto-positron",
        zoom=5
    )




    initial_fig.update_layout(
        title='Choropleth Map',
        margin={"r": 0, "t": 0, "l": 0, "b": 0},
        height=900,
        width=650
    )
   
    initial_fig.update_traces(hovertemplate='<b>%{hovertext}</b><br>CMCI Score: %{customdata}'
   
    )




    lon_manila = ph.loc[ph['PROVINCE'] == "Metro Manila", 'geometry'].get_coordinates().iloc[0]['x']
    lat_manila = ph.loc[ph['PROVINCE'] == "Metro Manila", 'geometry'].get_coordinates().iloc[0]['y']




    initial_fig.add_scattermapbox(
                lat=[lat_manila],
                lon=[lon_manila],
                mode='markers',
                text="Coordinates",
                marker_size=25,
                opacity=0.8,
                marker_color='rgb(235, 0, 100)',
                showlegend=False,
                name=""
            )




    if province:
        lon_selected = ph.loc[ph['PROVINCE'] == province, 'geometry'].get_coordinates().iloc[0]['x']
        lat_selected = ph.loc[ph['PROVINCE'] == province, 'geometry'].get_coordinates().iloc[0]['y']




        initial_fig.add_scattermapbox(
                lat=[lat_manila],
                lon=[lon_manila],
                mode='markers',
                text="Coordinates",
                marker_size=25,
                opacity=0.8,
                marker_color='rgb(235, 0, 100)',
                showlegend=False,
                name=""
            )
        initial_fig.add_scattermapbox(
                lat=[lat_manila, lat_selected],
                lon=[lon_manila, lon_selected],
                mode='markers',
                text="Coordinates",
                marker_size=25,
                opacity=0.8,
                marker_color='rgb(235, 0, 100)',
                showlegend=False,
                name=""
            )
       
    initial_fig.add_trace(initial_fig.data[0])




    return initial_fig




# Map 2
@app.callback(
    Output('choropleth-map', 'figure'),
    Input('map-year-dropdown-province', 'value'),
    Input('province-dropdown', 'value')
)
def update_choropleth(map_year, province):
    initial_column_values = p_choro.set_index('PROVINCE')[str(map_year)].replace('-', np.nan).astype(float).fillna(0)


    initial_fig = px.choropleth_mapbox(
        p_choro,
        geojson=p_choro,
        locations=p_choro.PROVINCE,
        featureidkey="properties.PROVINCE",
        color=initial_column_values,
        color_continuous_scale='Viridis',
        hover_name='PROVINCE',  
        hover_data={str(map_year): True},
        labels={map_year: 'Overall CMCI Score'},
        center={'lat': 12.8797, 'lon': 121.7740},
        mapbox_style="carto-positron",
        zoom=5
    )


    initial_fig.update_layout(
        title='Choropleth Map',
        margin={"r": 0, "t": 0, "l": 0, "b": 0},
        height=800,
    )
   
    initial_fig.update_traces(hovertemplate='<b>%{hovertext}</b><br>CMCI Score: %{customdata}'
   
    )


    lon_manila = ph.loc[ph['PROVINCE'] == "Metro Manila", 'geometry'].get_coordinates().iloc[0]['x']
    lat_manila = ph.loc[ph['PROVINCE'] == "Metro Manila", 'geometry'].get_coordinates().iloc[0]['y']


    initial_fig.add_scattermapbox(
                lat=[lat_manila],
                lon=[lon_manila],
                mode='markers',
                text="Coordinates",
                marker_size=15,
                opacity=0.8,
                marker_color='rgb(235, 0, 100)',
                showlegend=False,
                name=""
            )


    if province:
        lon_selected = ph.loc[ph['PROVINCE'] == province, 'geometry'].get_coordinates().iloc[0]['x']
        lat_selected = ph.loc[ph['PROVINCE'] == province, 'geometry'].get_coordinates().iloc[0]['y']


        initial_fig.add_scattermapbox(
            lat=[lat_manila],
            lon=[lon_manila],
            mode='markers',
            text="Manila",
            marker_size=15,
            opacity=0.8,
            marker_color='rgb(235, 0, 100)',
            showlegend=False,
            name=""
        )


        initial_fig.add_scattermapbox(
            lat=[lat_manila, lat_selected],
            lon=[lon_manila, lon_selected],
            mode='markers+lines',
            text=["Manila", province],
            marker_size=15,
            opacity=0.8,
            marker_color='rgb(235, 0, 100)',
            showlegend=False,
            name=""
        )


    initial_fig.add_trace(initial_fig.data[0])


    return initial_fig




@app.callback(Output('page-content', 'children'),
              [Input('url', 'pathname')])
def display_page(pathname):
    if pathname == '/' or pathname == '/page-1':
        return page1_layout
    elif pathname == '/page-2':
        return page2_layout
    elif pathname == '/page-3':
        return page3_layout
    else:
        return '404 - Page not found'




app.layout = html.Div([
    dcc.Location(id='url', refresh=False),
    navbar,
    html.Div(id='page-content')
])




if __name__ == '__main__':
    app.run_server(debug=False)

